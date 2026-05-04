import asyncio
import os
import re
import sqlite3
import sys
import time
import threading
import ipaddress
from collections import Counter
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import urlencode

from authlib.integrations.starlette_client import OAuth
from fastapi import FastAPI, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from sse_starlette.sse import EventSourceResponse
from openpyxl import load_workbook
from report import generate_report
from ai import build_assistant_reply
from reciever import receive_logs_handler
from event_bus import bus as event_bus
from alerts_cache import cache as alerts_cache
from logger import get_logger


def load_env_file() -> None:
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        try:
            with open(env_path, "w", encoding="utf-8") as env_file:
                env_file.write('SECRET_KEY="cyber_defense_ultra_secret_key_2025"\n')
                env_file.write('FRONTEND_URL="https://cyberdefensex.dpdns.org, https://cyberhub2025.github.io, http://localhost:3000, http://127.0.0.1:3000"\n')
                env_file.write('BACKEND_URL="http://localhost:8000"\n')
                env_file.write("NVIDIA_API_KEY=\n")
                env_file.write("GOOGLE_CLIENT_ID=\n")
                env_file.write("GOOGLE_CLIENT_SECRET=\n")
                env_file.write("GITHUB_CLIENT_ID=\n")
                env_file.write("GITHUB_CLIENT_SECRET=\n")
        except Exception as e:
            print(f"Warning: Could not create .env file automatically: {e}")
        return

    with open(env_path, "r", encoding="utf-8") as env_file:
        for raw_line in env_file:
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue

            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


load_env_file()


SECRET_KEY = os.environ.get("SECRET_KEY", "your_secret_key")
FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000")
BACKEND_URL = os.environ.get("BACKEND_URL", "http://localhost:8000")
ALERTS_XLSX_PATH = os.environ.get(
    "ALERTS_XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "Blockchain", "leader", "alerts.xlsx"),
)
APP_DATA_DB_PATH = os.environ.get("APP_DATA_DB_PATH", os.path.join(os.path.dirname(__file__), "app_data.db"))
REPORTS_DIR = os.environ.get(
    "REPORTS_DIR",
    os.path.join(os.path.dirname(__file__), "reports"),
)
ALLOWED_ALERT_STATUSES = {
    "active",
    "investigating",
    "blocked",
    "resolved",
    "mitigated",
}
DEFAULT_ALERT_STATUS = os.environ.get("DEFAULT_ALERT_STATUS", "active").strip().lower()
if DEFAULT_ALERT_STATUS not in ALLOWED_ALERT_STATUSES:
    DEFAULT_ALERT_STATUS = "active"

logger = get_logger("app")


# ---------------------------------------------------------------------------
# Lifespan (replaces deprecated @app.on_event)
# ---------------------------------------------------------------------------
@asynccontextmanager
async def lifespan(application: FastAPI):
    """Modern lifespan context manager for startup and shutdown."""
    # --- Startup ---
    event_bus.bind_loop(asyncio.get_running_loop())

    # Wire the alerts cache with the functions it needs
    alerts_cache.configure(
        sync_fn=sync_alerts_db_from_excel,
        load_fn=load_alerts_from_db,
        fetch_statuses_fn=fetch_statuses,
        insert_missing_fn=insert_missing_statuses,
        get_conn_fn=get_app_data_db_connection,
        default_status=DEFAULT_ALERT_STATUS,
        overview_builder=build_overview_payload,
        xlsx_path=ALERTS_XLSX_PATH,
    )

    # Initial cache warm-up
    try:
        alerts_cache.get_alerts()
        logger.info("Alerts cache warmed up successfully")
    except FileNotFoundError:
        logger.warning("Alerts xlsx not found during startup — cache empty")
    except Exception as exc:
        logger.error("Failed to warm alerts cache: %s", exc)

    # Subscribe to events that should invalidate the cache
    async def _cache_invalidator():
        async for _event in event_bus.subscribe("alerts.changed"):
            alerts_cache.invalidate()
            try:
                alerts_cache.get_alerts()
            except Exception as exc:
                logger.error("Cache refresh on event failed: %s", exc)

    invalidator_task = asyncio.create_task(_cache_invalidator())

    # Bootstrap blockchain workers (sub-app startup events don't fire automatically)
    _leader_module.worker_manager.bootstrap()
    # RLE stream monitor (background thread)
    start_rle_thread()
    # Broadcast watcher (background thread – waits briefly for server)
    start_broadcast_thread()
    # File watcher — real-time xlsx change detection
    start_file_watcher_thread()
    # Blockchain integrity checker (creates notifications on tamper)
    start_integrity_checker_thread()

    logger.info("Application startup complete")
    yield
    # --- Shutdown ---
    _shutdown_event.set()
    invalidator_task.cancel()
    logger.info("Application shutdown complete")


app = FastAPI(title="Cyber Defense Backend", lifespan=lifespan)

# ---------------------------------------------------------------------------
# Mount Blockchain Leader as a sub-application at /blockchain
# ---------------------------------------------------------------------------
import importlib
import importlib.util

LEADER_DIR = os.path.join(os.path.dirname(__file__), "Blockchain", "leader")

# Dynamically import the leader blockchain module from a non-package path
_leader_spec = importlib.util.spec_from_file_location(
    "blockchain_leader",
    os.path.join(LEADER_DIR, "blockchain.py"),
    submodule_search_locations=[LEADER_DIR],
)
_leader_module = importlib.util.module_from_spec(_leader_spec)
# Inject the leader dir into sys.path temporarily so config.py can be found
if LEADER_DIR not in sys.path:
    sys.path.insert(0, LEADER_DIR)
_leader_spec.loader.exec_module(_leader_module)

leader_app = _leader_module.app  # the FastAPI instance from blockchain.py
app.mount("/blockchain", leader_app)

# ---------------------------------------------------------------------------
# Background service state
# ---------------------------------------------------------------------------
_rle_thread = None
_broadcast_thread = None
_integrity_checker_thread = None
_shutdown_event = threading.Event()


# ---------------------------------------------------------------------------
# Background service helpers (RLE + Broadcast run as threads, not subprocesses)
# ---------------------------------------------------------------------------

def _run_rle_in_thread() -> None:
    """Import and run the RLE stream monitor in a daemon thread."""
    try:
        from rle import run_stream_monitor
        run_stream_monitor()
    except Exception as exc:
        logger.error("RLE thread error: %s", exc)


def _run_broadcast_in_thread() -> None:
    """Import and run the broadcast watcher in a daemon thread."""
    try:
        # Small delay to let the server start accepting requests
        time.sleep(3)
        broadcast_spec = importlib.util.spec_from_file_location(
            "broadcast",
            os.path.join(LEADER_DIR, "broadcast.py"),
        )
        broadcast_module = importlib.util.module_from_spec(broadcast_spec)
        broadcast_spec.loader.exec_module(broadcast_module)
        # Run the watch-and-broadcast loop
        broadcast_module.watch_and_broadcast(
            broadcast_module.DEFAULT_EXCEL_PATH,
            max(1, 8),
        )
    except Exception as exc:
        logger.error("Broadcast thread error: %s", exc)


def start_rle_thread() -> None:
    global _rle_thread
    if _rle_thread is not None and _rle_thread.is_alive():
        return
    _rle_thread = threading.Thread(target=_run_rle_in_thread, daemon=True)
    _rle_thread.start()


def start_broadcast_thread() -> None:
    global _broadcast_thread
    if _broadcast_thread is not None and _broadcast_thread.is_alive():
        return
    _broadcast_thread = threading.Thread(target=_run_broadcast_in_thread, daemon=True)
    _broadcast_thread.start()


# ---------------------------------------------------------------------------
# File watcher — monitors alerts.xlsx for ANY change and pushes via SSE
# ---------------------------------------------------------------------------
_file_watcher_thread = None
FILE_WATCH_INTERVAL = int(os.environ.get("FILE_WATCH_INTERVAL", "2"))  # seconds


def _run_file_watcher() -> None:
    """Poll alerts.xlsx mtime every N seconds. When it changes, emit an event
    so the cache refreshes and SSE pushes to all connected frontends."""
    last_mtime = 0.0
    try:
        if os.path.exists(ALERTS_XLSX_PATH):
            last_mtime = os.path.getmtime(ALERTS_XLSX_PATH)
    except OSError:
        pass

    logger.info("File watcher started — monitoring %s every %ds", ALERTS_XLSX_PATH, FILE_WATCH_INTERVAL)

    while not _shutdown_event.is_set():
        try:
            if os.path.exists(ALERTS_XLSX_PATH):
                current_mtime = os.path.getmtime(ALERTS_XLSX_PATH)
                if current_mtime != last_mtime:
                    last_mtime = current_mtime
                    logger.info("File change detected: %s", ALERTS_XLSX_PATH)
                    alerts_cache.invalidate()
                    event_bus.emit_threadsafe("alerts.changed", {"trigger": "file_watcher"})
        except OSError:
            pass
        except Exception as exc:
            logger.error("File watcher error: %s", exc)

        _shutdown_event.wait(FILE_WATCH_INTERVAL)


def start_file_watcher_thread() -> None:
    global _file_watcher_thread
    if _file_watcher_thread is not None and _file_watcher_thread.is_alive():
        return
    _file_watcher_thread = threading.Thread(target=_run_file_watcher, daemon=True)
    _file_watcher_thread.start()


INTEGRITY_CHECK_INTERVAL = int(os.environ.get("INTEGRITY_CHECK_INTERVAL", "10"))  # seconds


def _create_notification(severity: str, message: str, source: str = "blockchain") -> None:
    """Insert a notification row into app_data.db and emit an SSE event."""
    conn = get_app_data_db_connection()
    try:
        now = datetime.utcnow().isoformat() + "Z"
        conn.execute(
            "INSERT INTO notifications (severity, message, source, created_at) VALUES (?, ?, ?, ?)",
            (severity, message, source, now),
        )
        conn.commit()
    except Exception as exc:
        logger.error("Failed to insert notification: %s", exc)
    finally:
        conn.close()
    event_bus.emit_threadsafe("notifications.new", {"severity": severity, "message": message})


def _run_integrity_checker() -> None:
    """Periodically verify blockchain integrity and excel tampering.
    When an issue is detected, a notification is created."""
    import requests as _req

    # Wait for server startup
    time.sleep(8)
    logger.info("Integrity checker started — polling every %ds", INTEGRITY_CHECK_INTERVAL)

    last_state = {"blockchain_valid": True, "excel_intact": True, "excel_matches_chain": True}

    while not _shutdown_event.is_set():
        try:
            resp = _req.get(f"{BACKEND_URL}/blockchain/verify", timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                blockchain_valid = data.get("blockchain_valid", True)
                excel_intact = data.get("excel_intact", True)
                excel_matches = data.get("excel_matches_chain", True)
                integrity_ok = data.get("integrity_ok", True)

                # Blockchain chain validation failed
                if not blockchain_valid and last_state["blockchain_valid"]:
                    _create_notification(
                        "critical",
                        "Blockchain integrity violation detected — chain validation failed. Possible tampering.",
                        "blockchain",
                    )

                # Excel file was tampered
                if not excel_intact and last_state["excel_intact"]:
                    _create_notification(
                        "critical",
                        "Alert data (alerts.xlsx) has been tampered with. The file does not match the trusted backup.",
                        "blockchain",
                    )

                # Excel doesn't match latest block hash
                if not excel_matches and last_state["excel_matches_chain"]:
                    _create_notification(
                        "high",
                        "Excel file hash does not match the latest blockchain block. Data may have been modified outside the system.",
                        "blockchain",
                    )

                # Recovery detected
                if integrity_ok and (not last_state["blockchain_valid"] or not last_state["excel_intact"]):
                    _create_notification(
                        "success",
                        "Blockchain integrity has been restored. System is operating normally.",
                        "blockchain",
                    )

                last_state = {
                    "blockchain_valid": blockchain_valid,
                    "excel_intact": excel_intact,
                    "excel_matches_chain": excel_matches,
                }
        except Exception as exc:
            logger.debug("Integrity checker poll error: %s", exc)

        _shutdown_event.wait(INTEGRITY_CHECK_INTERVAL)


def start_integrity_checker_thread() -> None:
    global _integrity_checker_thread
    if _integrity_checker_thread is not None and _integrity_checker_thread.is_alive():
        return
    _integrity_checker_thread = threading.Thread(target=_run_integrity_checker, daemon=True)
    _integrity_checker_thread.start()

# Split FRONTEND_URL by comma to allow multiple origins (e.g. localhost, github pages)
ALLOWED_ORIGINS = [url.strip() for url in FRONTEND_URL.split(",") if url.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)


oauth = OAuth()
oauth.register(
    name="google",
    client_id=os.environ.get("GOOGLE_CLIENT_ID"),
    client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)
oauth.register(
    name="github",
    client_id=os.environ.get("GITHUB_CLIENT_ID"),
    client_secret=os.environ.get("GITHUB_CLIENT_SECRET"),
    access_token_url="https://github.com/login/oauth/access_token",
    authorize_url="https://github.com/login/oauth/authorize",
    api_base_url="https://api.github.com/",
    client_kwargs={"scope": "user:email"},
)


def init_db() -> None:
    conn = sqlite3.connect("users.db")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            fullname TEXT,
            username TEXT UNIQUE,
            password TEXT,
            oauth_provider TEXT,
            oauth_id TEXT,
            job_title TEXT,
            department TEXT,
            timezone TEXT,
            avatar TEXT,
            appearance TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def init_app_data_db() -> None:
    """Create all non-user tables in the unified app_data.db."""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    conn = sqlite3.connect(APP_DATA_DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS alerts (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            type TEXT NOT NULL,
            severity TEXT NOT NULL,
            source TEXT NOT NULL,
            target TEXT NOT NULL,
            detected_at TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_alerts_detected_at
        ON alerts(detected_at)
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_alerts_severity
        ON alerts(severity)
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS threat_status (
            threat_id TEXT PRIMARY KEY,
            status TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS report_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id TEXT UNIQUE,
            generated_at TEXT NOT NULL,
            file_name TEXT NOT NULL,
            file_path TEXT NOT NULL,
            row_count INTEGER NOT NULL,
            file_size INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS senders (
            ip TEXT PRIMARY KEY,
            hostname TEXT,
            total_logs INTEGER NOT NULL DEFAULT 0,
            first_seen TEXT NOT NULL,
            last_seen TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            severity TEXT NOT NULL DEFAULT 'info',
            message TEXT NOT NULL,
            source TEXT NOT NULL DEFAULT 'system',
            created_at TEXT NOT NULL,
            read INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_notifications_created_at
        ON notifications(created_at DESC)
        """
    )
    conn.commit()
    conn.close()


init_db()
init_app_data_db()


def get_db_connection() -> sqlite3.Connection:
    conn = sqlite3.connect("users.db")
    conn.row_factory = sqlite3.Row
    return conn


def get_app_data_db_connection() -> sqlite3.Connection:
    """Return a connection to the unified app_data.db."""
    conn = sqlite3.connect(APP_DATA_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def normalize_header(header: Any) -> str:
    if header is None:
        return ""

    text = str(header).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def format_cell_value(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()
    return text or None


def normalize_severity(value: Any) -> str:
    text = (format_cell_value(value) or "").lower()
    if "critical" in text:
        return "critical"
    if "high" in text:
        return "high"
    if "medium" in text:
        return "medium"
    if "low" in text:
        return "low"
    return "medium"


def coerce_status(value: Optional[str]) -> str:
    status = (value or "").strip().lower()
    if status in ALLOWED_ALERT_STATUSES:
        return status
    return DEFAULT_ALERT_STATUS


def pick_value(raw: Dict[str, Any], keys: Iterable[str]) -> Any:
    for key in keys:
        if key in raw and raw[key] not in (None, ""):
            return raw[key]
    return None


def load_alerts_from_excel() -> List[Dict[str, str]]:
    if not os.path.exists(ALERTS_XLSX_PATH):
        raise FileNotFoundError(ALERTS_XLSX_PATH)

    workbook = load_workbook(ALERTS_XLSX_PATH, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []

        headers = [normalize_header(header) for header in header_row]
        alerts: List[Dict[str, str]] = []

        id_keys = ("threat_id", "threatid", "alert_id", "alertid", "id")
        name_keys = ("name", "attack", "threat_name", "alert_name", "title")
        type_keys = ("type", "threat_type", "alert_type", "category")
        severity_keys = ("severity", "risk", "priority")
        source_keys = (
            "source",
            "source_ip",
            "sourceip",
            "src",
            "src_ip",
            "source_address",
        )
        target_keys = (
            "target",
            "target_host",
            "destination",
            "dest",
            "dst",
            "dst_ip",
            "target_ip",
        )
        detected_keys = (
            "detected_at",
            "detected",
            "timestamp",
            "time",
            "date",
            "detected_time",
            "detectedat",
            "start_time",
            "starttime",
        )

        for row_index, row in enumerate(rows, start=2):
            if not row:
                continue

            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            raw: Dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                raw[header] = row[idx] if idx < len(row) else None

            threat_id_value = pick_value(raw, id_keys)
            threat_id = format_cell_value(threat_id_value)
            if not threat_id:
                threat_id = f"ALERT-{row_index - 1:04d}"

            alert = {
                "id": threat_id,
                "name": format_cell_value(pick_value(raw, name_keys)) or "Unknown",
                "type": format_cell_value(pick_value(raw, type_keys)) or "Unknown",
                "severity": normalize_severity(pick_value(raw, severity_keys)),
                "source": format_cell_value(pick_value(raw, source_keys)) or "-",
                "target": format_cell_value(pick_value(raw, target_keys)) or "-",
                "detectedAt": format_cell_value(pick_value(raw, detected_keys)) or "-",
            }
            alerts.append(alert)

        return alerts
    finally:
        workbook.close()


def sync_alerts_db_from_excel() -> int:
    alerts = load_alerts_from_excel()
    conn = get_app_data_db_connection()
    try:
        now = datetime.utcnow().isoformat() + "Z"
        ordered_ids: List[str] = []

        for row_index, alert in enumerate(alerts, start=1):
            ordered_ids.append(alert["id"])
            conn.execute(
                """
                INSERT INTO alerts (id, name, type, severity, source, target, detected_at, row_index, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(id)
                DO UPDATE SET
                    name=excluded.name,
                    type=excluded.type,
                    severity=excluded.severity,
                    source=excluded.source,
                    target=excluded.target,
                    detected_at=excluded.detected_at,
                    row_index=excluded.row_index,
                    updated_at=excluded.updated_at
                """,
                (
                    alert["id"],
                    alert["name"],
                    alert["type"],
                    alert["severity"],
                    alert["source"],
                    alert["target"],
                    alert["detectedAt"],
                    row_index,
                    now,
                ),
            )

        if ordered_ids:
            placeholders = ",".join(["?"] * len(ordered_ids))
            conn.execute(f"DELETE FROM alerts WHERE id NOT IN ({placeholders})", ordered_ids)
        else:
            conn.execute("DELETE FROM alerts")

        conn.commit()
        return len(alerts)
    finally:
        conn.close()


def load_alerts_from_db() -> List[Dict[str, str]]:
    conn = get_app_data_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, name, type, severity, source, target, detected_at
            FROM alerts
            ORDER BY row_index ASC
            """
        ).fetchall()

        return [
            {
                "id": row["id"],
                "name": row["name"],
                "type": row["type"],
                "severity": row["severity"],
                "source": row["source"],
                "target": row["target"],
                "detectedAt": row["detected_at"],
            }
            for row in rows
        ]
    finally:
        conn.close()


def get_alerts_with_status() -> List[Dict[str, str]]:
    sync_alerts_db_from_excel()
    alerts = load_alerts_from_db()

    conn = get_app_data_db_connection()
    try:
        threat_ids = [alert["id"] for alert in alerts]
        status_map = fetch_statuses(conn, threat_ids)
        insert_missing_statuses(conn, threat_ids, status_map)
        conn.commit()
    finally:
        conn.close()

    for alert in alerts:
        alert["status"] = status_map.get(alert["id"], DEFAULT_ALERT_STATUS)

    return alerts


def parse_detected_at(value: Optional[str]) -> Optional[datetime]:
    text = (value or "").strip()
    if not text or text == "-":
        return None

    normalized = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(normalized).replace(tzinfo=None)
    except ValueError:
        pass

    patterns = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    )
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def calculate_change(current: int, previous: int) -> Dict[str, str]:
    if previous <= 0:
        if current <= 0:
            return {"change": "0%", "trend": "up"}
        return {"change": "+100%", "trend": "up"}

    percent = int(round(((current - previous) / previous) * 100))
    prefix = "+" if percent > 0 else ""
    trend = "up" if percent >= 0 else "down"
    return {"change": f"{prefix}{percent}%", "trend": trend}


def infer_origin(source: str) -> str:
    text = (source or "").strip()
    lower = text.lower()

    keyword_map = {
        "russia": "Russia",
        "china": "China",
        "north korea": "North Korea",
        "iran": "Iran",
        "usa": "United States",
        "united states": "United States",
        "germany": "Germany",
        "india": "India",
        "japan": "Japan",
        "brazil": "Brazil",
    }
    for keyword, country in keyword_map.items():
        if keyword in lower:
            return country

    cleaned = text.split(":")[0]
    try:
        ip = ipaddress.ip_address(cleaned)
        if ip.is_private:
            return "Internal Network"

        if isinstance(ip, ipaddress.IPv4Address):
            country_pool = [
                "United States",
                "Germany",
                "Singapore",
                "Brazil",
                "India",
                "Japan",
            ]
            return country_pool[int(str(ip).split(".")[0]) % len(country_pool)]
    except ValueError:
        pass

    return "Unknown"


def to_relative_time(value: Optional[datetime]) -> str:
    if not value:
        return "Unknown"

    delta = datetime.utcnow() - value
    seconds = int(max(delta.total_seconds(), 0))
    if seconds < 60:
        return f"{seconds}s ago"
    minutes = seconds // 60
    if minutes < 60:
        return f"{minutes} min ago"
    hours = minutes // 60
    if hours < 24:
        return f"{hours}h ago"
    days = hours // 24
    return f"{days}d ago"


def build_overview_payload(alerts: List[Dict[str, str]]) -> Dict[str, Any]:
    now = datetime.utcnow()
    parsed_rows: List[Dict[str, Any]] = []
    for alert in alerts:
        detected_dt = parse_detected_at(alert.get("detectedAt"))
        parsed_rows.append({**alert, "detected_dt": detected_dt})

    current_24h_start = now - timedelta(hours=24)
    previous_24h_start = now - timedelta(hours=48)
    alerts_current_24h = [a for a in parsed_rows if a["detected_dt"] and a["detected_dt"] >= current_24h_start]
    alerts_previous_24h = [
        a
        for a in parsed_rows
        if a["detected_dt"] and previous_24h_start <= a["detected_dt"] < current_24h_start
    ]

    active_now = [a for a in parsed_rows if a.get("status") in {"active", "investigating"}]
    active_prev = [a for a in alerts_previous_24h if a.get("status") in {"active", "investigating"}]
    vuln_now = [a for a in parsed_rows if a.get("severity") in {"critical", "high"}]
    vuln_prev = [a for a in alerts_previous_24h if a.get("severity") in {"critical", "high"}]

    current_assets = {a.get("target") for a in parsed_rows if a.get("target") and a.get("target") != "-"}
    prev_assets = {a.get("target") for a in alerts_previous_24h if a.get("target") and a.get("target") != "-"}

    active_change = calculate_change(len(active_now), len(active_prev))
    alerts_change = calculate_change(len(alerts_current_24h), len(alerts_previous_24h))
    vuln_change = calculate_change(len(vuln_now), len(vuln_prev))
    asset_change = calculate_change(len(current_assets), len(prev_assets))

    severity_counter = Counter(a.get("severity", "medium") for a in parsed_rows)
    severity_data = [
        {"name": "Critical", "value": severity_counter.get("critical", 0), "color": "#ef4444"},
        {"name": "High", "value": severity_counter.get("high", 0), "color": "#f97316"},
        {"name": "Medium", "value": severity_counter.get("medium", 0), "color": "#f59e0b"},
        {"name": "Low", "value": severity_counter.get("low", 0), "color": "#10b981"},
    ]

    attack_name_counter = Counter(a.get("name") or "Unknown" for a in parsed_rows)
    attack_types_data = [
        {"type": threat_name, "count": count}
        for threat_name, count in attack_name_counter.most_common(10)
    ]

    threat_activity: List[Dict[str, int]] = []
    for day_offset in range(6, -1, -1):
        day_start = (now - timedelta(days=day_offset)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)
        day_alerts = [a for a in parsed_rows if a["detected_dt"] and day_start <= a["detected_dt"] < day_end]
        blocked_count = sum(1 for a in day_alerts if a.get("status") in {"blocked", "resolved", "mitigated"})
        threat_activity.append(
            {
                "name": day_start.strftime("%a"),
                "threats": len(day_alerts),
                "blocked": blocked_count,
            }
        )

    network_traffic: List[Dict[str, int]] = []
    for hour in range(0, 24, 4):
        bucket_alerts = [
            a
            for a in parsed_rows
            if a["detected_dt"] and a["detected_dt"].hour >= hour and a["detected_dt"].hour < hour + 4
        ]
        blocked_bucket = sum(1 for a in bucket_alerts if a.get("status") in {"blocked", "resolved", "mitigated"})
        inbound = len(bucket_alerts) * 20 + 60
        outbound = blocked_bucket * 18 + 40
        network_traffic.append(
            {
                "time": f"{hour:02d}:00",
                "inbound": inbound,
                "outbound": outbound,
            }
        )

    def _extract_id_number(alert):
        """Extract the numeric portion of a threat ID for sorting (e.g. 'ALERT-0042' -> 42)."""
        raw = alert.get("id") or ""
        # Try to pull digits from the end (handles 'ALERT-0042', 'THR-123', etc.)
        match = re.search(r"(\d+)$", raw)
        if match:
            return int(match.group(1))
        # Fallback: try converting the whole thing
        try:
            return int(raw)
        except (ValueError, TypeError):
            return 0

    sorted_alerts = sorted(
        parsed_rows,
        key=_extract_id_number,
        reverse=True,
    )
    recent_alerts = [
        {
            "id": alert.get("id"),
            "type": alert.get("severity", "medium"),
            "message": f"{alert.get('name', 'Unknown')} detected ({alert.get('type', 'Unknown')})",
            "time": to_relative_time(alert.get("detected_dt")),
            "source": alert.get("source", "-"),
        }
        for alert in sorted_alerts[:5]
    ]

    origin_counter = Counter(
        alert.get("source", "-") for alert in parsed_rows if alert.get("source") and alert.get("source") != "-"
    )
    threat_origins = [
        {"ip": ip, "count": count}
        for ip, count in origin_counter.most_common(10)
    ]

    return {
        "stats": {
            "activeThreats": {
                "value": len(active_now),
                "change": active_change["change"],
                "trend": active_change["trend"],
            },
            "alertsToday": {
                "value": len(alerts_current_24h),
                "change": alerts_change["change"],
                "trend": alerts_change["trend"],
            },
            "vulnerabilities": {
                "value": len(vuln_now),
                "change": vuln_change["change"],
                "trend": vuln_change["trend"],
            },
            "protectedAssets": {
                "value": len(current_assets),
                "change": asset_change["change"],
                "trend": asset_change["trend"],
            },
        },
        "threatActivity": threat_activity,
        "severityDistribution": severity_data,
        "attackTypes": attack_types_data,
        "networkTraffic": network_traffic,
        "recentAlerts": recent_alerts,
        "threatOrigins": threat_origins,
    }


def build_report_id(timestamp: datetime) -> str:
    return f"RPT-{timestamp.strftime('%Y%m%d%H%M%S')}"


def fetch_statuses(conn: sqlite3.Connection, threat_ids: List[str]) -> Dict[str, str]:
    if not threat_ids:
        return {}

    placeholders = ",".join(["?"] * len(threat_ids))
    rows = conn.execute(
        f"SELECT threat_id, status FROM threat_status WHERE threat_id IN ({placeholders})",
        threat_ids,
    ).fetchall()
    return {row["threat_id"]: coerce_status(row["status"]) for row in rows}


def insert_missing_statuses(
    conn: sqlite3.Connection,
    threat_ids: List[str],
    existing: Dict[str, str],
) -> None:
    missing = [threat_id for threat_id in threat_ids if threat_id not in existing]
    if not missing:
        return

    now = datetime.utcnow().isoformat() + "Z"
    conn.executemany(
        "INSERT OR IGNORE INTO threat_status (threat_id, status, updated_at) VALUES (?, ?, ?)",
        [(threat_id, DEFAULT_ALERT_STATUS, now) for threat_id in missing],
    )
    for threat_id in missing:
        existing[threat_id] = DEFAULT_ALERT_STATUS


@app.get("/", response_class=HTMLResponse)
async def index() -> str:
    return "<h1>Backend is running</h1>"


@app.get("/signup", response_class=HTMLResponse)
async def signup_page() -> str:
    return "<h1>Signup endpoint is available</h1>"


@app.post("/signup")
async def signup(request: Request):
    form = await request.form()
    fullname = (form.get("fullname") or "").strip()
    username = (form.get("username") or "").strip()
    password = (form.get("password") or "")

    if not fullname or not username or not password:
        return JSONResponse(
            {"success": False, "message": "All fields are required"},
            status_code=400,
        )

    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO users (fullname, username, password) VALUES (?, ?, ?)",
            (fullname, username, password),
        )
        conn.commit()
        conn.close()
        return RedirectResponse(url="/login", status_code=302)
    except sqlite3.IntegrityError:
        return JSONResponse(
            {"success": False, "message": "Username already exists"},
            status_code=409,
        )


@app.get("/login", response_class=HTMLResponse)
async def login_page() -> str:
    return "<h1>Login endpoint is available</h1>"


@app.post("/login")
async def login(request: Request):
    form = await request.form()
    username = (form.get("username") or "").strip()
    password = (form.get("password") or "")

    if not username or not password:
        return JSONResponse(
            {"success": False, "message": "Username and password are required"},
            status_code=400,
        )

    conn = get_db_connection()
    user = conn.execute(
        "SELECT * FROM users WHERE username=? AND password=?",
        (username, password),
    ).fetchone()
    conn.close()

    if user:
        request.session["username"] = user["username"]
        request.session["fullname"] = user["fullname"]
        return RedirectResponse(url="/home", status_code=302)

    return JSONResponse(
        {"success": False, "message": "Invalid credentials"},
        status_code=401,
    )


@app.post("/api/signup")
async def api_signup(request: Request):
    data = await request.json()
    fullname = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""

    if not fullname or not email or not password:
        return JSONResponse(
            {"success": False, "message": "Name, email, and password are required"},
            status_code=400,
        )

    conn = None
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO users (fullname, username, password) VALUES (?, ?, ?)",
            (fullname, email, password),
        )
        conn.commit()
        return JSONResponse(
            {"success": True, "message": "Account created successfully"},
            status_code=201,
        )
    except sqlite3.IntegrityError:
        return JSONResponse(
            {
                "success": False,
                "message": "An account with this email already exists. Please login.",
            },
            status_code=409,
        )
    except Exception:
        return JSONResponse(
            {"success": False, "message": "Internal server error"},
            status_code=500,
        )
    finally:
        if conn:
            conn.close()


@app.post("/api/login")
async def api_login(request: Request):
    data = await request.json()
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""

    if not email or not password:
        return JSONResponse(
            {"success": False, "message": "Email and password are required"},
            status_code=400,
        )

    conn = None
    try:
        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE username=? AND password=?",
            (email, password),
        ).fetchone()

        if not user:
            return JSONResponse(
                {
                    "success": False,
                    "message": "Incorrect email or password. Please try again.",
                },
                status_code=401,
            )

        request.session["username"] = user["username"]
        request.session["fullname"] = user["fullname"]

        return JSONResponse(
            {
                "success": True,
                "user": {"name": user["fullname"], "email": user["username"]},
            }
        )
    except Exception:
        return JSONResponse(
            {"success": False, "message": "Internal server error"},
            status_code=500,
        )
    finally:
        if conn:
            conn.close()


@app.get("/home")
async def home(request: Request):
    if "username" in request.session:
        return JSONResponse(
            {
                "username": request.session.get("username"),
                "fullname": request.session.get("fullname"),
            }
        )
    return RedirectResponse(url="/login", status_code=302)


@app.get("/api/user/profile")
async def get_user_profile(request: Request):
    username = request.session.get("username")
    if not username:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
        
    conn = get_db_connection()
    try:
        user = conn.execute("SELECT fullname, username, job_title, department, timezone, avatar, appearance FROM users WHERE username=?", (username,)).fetchone()
        if not user:
            return JSONResponse({"success": False, "message": "User not found"}, status_code=404)
            
        appearance_data = {}
        try:
            if user["appearance"]:
                import json
                appearance_data = json.loads(user["appearance"])
        except Exception:
            pass
            
        return JSONResponse({
            "success": True,
            "profile": {
                "fullName": user["fullname"] or "",
                "email": user["username"] or "",
                "jobTitle": user["job_title"] or "",
                "department": user["department"] or "",
                "timezone": user["timezone"] or "",
                "avatar": user["avatar"]
            },
            "appearance": appearance_data
        })
    finally:
        conn.close()


@app.put("/api/user/profile")
async def update_user_profile(request: Request):
    username = request.session.get("username")
    if not username:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
        
    data = await request.json()
    profile = data.get("profile", {})
    appearance = data.get("appearance", {})
    
    conn = get_db_connection()
    try:
        import json
        appearance_str = json.dumps(appearance) if appearance else None
        
        conn.execute("""
            UPDATE users 
            SET fullname=?, job_title=?, department=?, timezone=?, avatar=?, appearance=?
            WHERE username=?
        """, (
            profile.get("fullName"),
            profile.get("jobTitle"),
            profile.get("department"),
            profile.get("timezone"),
            profile.get("avatar"),
            appearance_str,
            username
        ))
        conn.commit()
        
        # Update session name just in case
        if profile.get("fullName"):
            request.session["fullname"] = profile.get("fullName")
            
        return JSONResponse({"success": True, "message": "Profile updated successfully"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)
    finally:
        conn.close()


@app.get("/logout")
async def logout(request: Request):
    request.session.pop("username", None)
    request.session.pop("fullname", None)
    return RedirectResponse(url="/login", status_code=302)


@app.get("/auth/google")
async def google_login(request: Request):
    redirect_uri = f"{BACKEND_URL}/auth/google/callback"
    return await oauth.google.authorize_redirect(request, redirect_uri)


@app.get("/auth/google/callback")
async def google_callback(request: Request):
    try:
        token = await oauth.google.authorize_access_token(request)
        user_info = token.get("userinfo")

        if not user_info:
            user_info = (await oauth.google.get("https://openidconnect.googleapis.com/v1/userinfo", token=token)).json()

        oauth_id = user_info.get("sub")
        email = user_info.get("email")
        name = user_info.get("name", email.split("@")[0])

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE oauth_provider=? AND oauth_id=?",
            ("google", oauth_id),
        ).fetchone()

        if not user:
            existing = conn.execute(
                "SELECT * FROM users WHERE username=?",
                (email,),
            ).fetchone()

            if existing:
                conn.execute(
                    "UPDATE users SET oauth_provider=?, oauth_id=? WHERE username=?",
                    ("google", oauth_id, email),
                )
                user = existing
            else:
                conn.execute(
                    "INSERT INTO users (fullname, username, oauth_provider, oauth_id) VALUES (?, ?, ?, ?)",
                    (name, email, "google", oauth_id),
                )
                user = conn.execute(
                    "SELECT * FROM users WHERE username=?",
                    (email,),
                ).fetchone()

        conn.commit()
        conn.close()

        request.session["username"] = user["username"]
        request.session["fullname"] = user["fullname"]

        params = urlencode(
            {
                "name": user["fullname"],
                "email": user["username"],
                "provider": "google",
            }
        )
        return RedirectResponse(url=f"{FRONTEND_URL}/oauth/success?{params}", status_code=302)
    except Exception as exc:
        params = urlencode({"error": str(exc)})
        return RedirectResponse(url=f"{FRONTEND_URL}/login?{params}", status_code=302)


@app.get("/api/auth/google/callback")
async def api_google_callback(request: Request):
    try:
        code = request.query_params.get("code")
        if not code:
            return JSONResponse(
                {"success": False, "message": "No authorization code provided"},
                status_code=400,
            )

        token = await oauth.google.authorize_access_token(request)
        user_info = token.get("userinfo")

        if not user_info:
            user_info = (await oauth.google.get("https://openidconnect.googleapis.com/v1/userinfo", token=token)).json()

        oauth_id = user_info.get("sub")
        email = user_info.get("email")
        name = user_info.get("name", email.split("@")[0])

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE oauth_provider=? AND oauth_id=?",
            ("google", oauth_id),
        ).fetchone()

        if not user:
            existing = conn.execute(
                "SELECT * FROM users WHERE username=?",
                (email,),
            ).fetchone()

            if existing:
                conn.execute(
                    "UPDATE users SET oauth_provider=?, oauth_id=? WHERE username=?",
                    ("google", oauth_id, email),
                )
                user = existing
            else:
                conn.execute(
                    "INSERT INTO users (fullname, username, oauth_provider, oauth_id) VALUES (?, ?, ?, ?)",
                    (name, email, "google", oauth_id),
                )
                user = conn.execute(
                    "SELECT * FROM users WHERE username=?",
                    (email,),
                ).fetchone()

        conn.commit()
        conn.close()

        request.session["username"] = user["username"]
        request.session["fullname"] = user["fullname"]

        return JSONResponse(
            {
                "success": True,
                "user": {"name": user["fullname"], "email": user["username"]},
            }
        )
    except Exception as exc:
        return JSONResponse(
            {"success": False, "message": str(exc)},
            status_code=500,
        )


@app.get("/auth/github")
async def github_login(request: Request):
    redirect_uri = f"{BACKEND_URL}/auth/github/callback"
    return await oauth.github.authorize_redirect(request, redirect_uri)


@app.get("/auth/github/callback")
async def github_callback(request: Request):
    try:
        token = await oauth.github.authorize_access_token(request)

        resp = await oauth.github.get("user", token=token)
        user_info = resp.json()

        oauth_id = str(user_info.get("id"))
        username = user_info.get("login")
        name = user_info.get("name") or username

        email_resp = await oauth.github.get("user/emails", token=token)
        emails = email_resp.json()
        email = next((e["email"] for e in emails if e.get("primary")), f"{username}@github.local")

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE oauth_provider=? AND oauth_id=?",
            ("github", oauth_id),
        ).fetchone()

        if not user:
            existing = conn.execute(
                "SELECT * FROM users WHERE username=?",
                (username,),
            ).fetchone()

            if existing:
                conn.execute(
                    "UPDATE users SET oauth_provider=?, oauth_id=? WHERE username=?",
                    ("github", oauth_id, username),
                )
                user = existing
            else:
                conn.execute(
                    "INSERT INTO users (fullname, username, oauth_provider, oauth_id) VALUES (?, ?, ?, ?)",
                    (name, username, "github", oauth_id),
                )
                user = conn.execute(
                    "SELECT * FROM users WHERE username=?",
                    (username,),
                ).fetchone()

        conn.commit()
        conn.close()

        request.session["username"] = user["username"]
        request.session["fullname"] = user["fullname"]

        params = urlencode(
            {
                "name": user["fullname"],
                "email": email,
                "provider": "github",
            }
        )
        return RedirectResponse(url=f"{FRONTEND_URL}/oauth/success?{params}", status_code=302)
    except Exception as exc:
        params = urlencode({"error": str(exc)})
        return RedirectResponse(url=f"{FRONTEND_URL}/login?{params}", status_code=302)


@app.get("/api/auth/github/callback")
async def api_github_callback(request: Request):
    try:
        code = request.query_params.get("code")
        if not code:
            return JSONResponse(
                {"success": False, "message": "No authorization code provided"},
                status_code=400,
            )

        token = await oauth.github.authorize_access_token(request)

        resp = await oauth.github.get("user", token=token)
        user_info = resp.json()

        oauth_id = str(user_info.get("id"))
        username = user_info.get("login")
        name = user_info.get("name") or username

        email_resp = await oauth.github.get("user/emails", token=token)
        emails = email_resp.json()
        email = next((e["email"] for e in emails if e.get("primary")), f"{username}@github.local")

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE oauth_provider=? AND oauth_id=?",
            ("github", oauth_id),
        ).fetchone()

        if not user:
            existing = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

            if existing:
                conn.execute(
                    "UPDATE users SET oauth_provider=?, oauth_id=? WHERE username=?",
                    ("github", oauth_id, username),
                )
                user = existing
            else:
                conn.execute(
                    "INSERT INTO users (fullname, username, oauth_provider, oauth_id) VALUES (?, ?, ?, ?)",
                    (name, username, "github", oauth_id),
                )
                user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

        conn.commit()
        conn.close()

        request.session["username"] = user["username"]
        request.session["fullname"] = user["fullname"]

        return JSONResponse(
            {
                "success": True,
                "user": {"name": user["fullname"], "email": email},
            }
        )
    except Exception as exc:
        return JSONResponse(
            {"success": False, "message": str(exc)},
            status_code=500,
        )


@app.get("/api/alerts")
async def api_get_alerts():
    try:
        alerts = alerts_cache.get_alerts()
    except FileNotFoundError:
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    return JSONResponse(
        {
            "alerts": alerts,
            "count": len(alerts),
            "updatedAt": datetime.utcnow().isoformat() + "Z",
        }
    )


@app.get("/api/overview")
async def api_get_overview():
    try:
        alerts = alerts_cache.get_alerts()
    except FileNotFoundError:
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    try:
        payload = alerts_cache.get_overview()
    except Exception:
        payload = build_overview_payload(alerts)
    payload["count"] = len(alerts)
    payload["updatedAt"] = datetime.utcnow().isoformat() + "Z"
    return JSONResponse(payload)


@app.post("/api/assistant/chat")
async def api_assistant_chat(request: Request):
    payload = await request.json()
    messages = payload.get("messages", [])

    if not isinstance(messages, list) or not messages:
        return JSONResponse(
            {"message": "messages must be a non-empty list"},
            status_code=400,
        )

    try:
        alerts = alerts_cache.get_alerts()
    except FileNotFoundError:
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    result = build_assistant_reply(messages, alerts)
    return JSONResponse(result)


@app.patch("/api/alerts/{threat_id}/status")
async def api_update_alert_status(threat_id: str, request: Request):
    payload = await request.json()
    status = (payload.get("status") or "").strip().lower()

    if status not in ALLOWED_ALERT_STATUSES:
        return JSONResponse(
            {
                "message": "Invalid status value.",
                "allowed": sorted(ALLOWED_ALERT_STATUSES),
            },
            status_code=400,
        )

    try:
        alerts = alerts_cache.get_alerts()
    except FileNotFoundError:
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    alert_ids = {alert["id"] for alert in alerts}
    if threat_id not in alert_ids:
        return JSONResponse(
            {"message": "Alert not found for provided threat id."},
            status_code=404,
        )

    conn = get_app_data_db_connection()
    try:
        now = datetime.utcnow().isoformat() + "Z"
        conn.execute(
            """
            INSERT INTO threat_status (threat_id, status, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(threat_id)
            DO UPDATE SET status=excluded.status, updated_at=excluded.updated_at
            """,
            (threat_id, status, now),
        )
        conn.commit()
    finally:
        conn.close()

    # Invalidate cache and notify connected frontends
    alerts_cache.invalidate()
    event_bus.emit_threadsafe("alerts.changed", {"trigger": "status_update", "threat_id": threat_id})

    return JSONResponse({"success": True, "threat_id": threat_id, "status": status})


# ---------------------------------------------------------------------------
# Notifications API
# ---------------------------------------------------------------------------
@app.get("/api/notifications")
async def api_get_notifications():
    """Return the latest notifications (most recent first)."""
    conn = get_app_data_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT id, severity, message, source, created_at, read
            FROM notifications
            ORDER BY created_at DESC
            LIMIT 50
            """
        ).fetchall()
    finally:
        conn.close()

    notifications = [
        {
            "id": row["id"],
            "severity": row["severity"],
            "message": row["message"],
            "source": row["source"],
            "time": to_relative_time(parse_detected_at(row["created_at"])),
            "createdAt": row["created_at"],
            "read": bool(row["read"]),
        }
        for row in rows
    ]

    return JSONResponse({"notifications": notifications, "count": len(notifications)})


@app.delete("/api/notifications/{notification_id}")
async def api_delete_notification(notification_id: int):
    """Remove a single notification."""
    conn = get_app_data_db_connection()
    try:
        cursor = conn.execute(
            "DELETE FROM notifications WHERE id = ?", (notification_id,)
        )
        conn.commit()
        deleted = cursor.rowcount > 0
    finally:
        conn.close()

    if not deleted:
        return JSONResponse({"message": "Notification not found."}, status_code=404)

    event_bus.emit_threadsafe("notifications.new", {"trigger": "deleted"})
    return JSONResponse({"success": True, "deleted_id": notification_id})


@app.delete("/api/notifications")
async def api_clear_all_notifications():
    """Remove all notifications."""
    conn = get_app_data_db_connection()
    try:
        cursor = conn.execute("DELETE FROM notifications")
        conn.commit()
        count = cursor.rowcount
    finally:
        conn.close()

    event_bus.emit_threadsafe("notifications.new", {"trigger": "cleared"})
    return JSONResponse({"success": True, "cleared": count})


@app.get("/api/reports")
async def api_list_reports():
    conn = get_app_data_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT report_id, generated_at, file_name, file_path, row_count, file_size
            FROM report_runs
            ORDER BY generated_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    reports = [
        {
            "id": row["report_id"],
            "generatedAt": row["generated_at"],
            "fileName": row["file_name"],
            "rowCount": row["row_count"],
            "sizeBytes": row["file_size"],
            "downloadUrl": f"/api/reports/{row['report_id']}/download",
        }
        for row in rows
    ]

    return JSONResponse({"reports": reports, "count": len(reports)})


@app.post("/api/reports/generate")
async def api_generate_report():
    if not os.path.exists(ALERTS_XLSX_PATH):
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )

    generated_at = datetime.utcnow()
    report_id = build_report_id(generated_at)
    file_name = f"security_alert_report_{generated_at.strftime('%Y%m%d_%H%M%S')}.pdf"

    try:
        result = generate_report(ALERTS_XLSX_PATH, REPORTS_DIR, file_name)
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    output_path = result["output_path"]
    file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0

    conn = get_app_data_db_connection()
    try:
        conn.execute(
            """
            INSERT INTO report_runs (report_id, generated_at, file_name, file_path, row_count, file_size)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                report_id,
                generated_at.isoformat() + "Z",
                file_name,
                output_path,
                result["row_count"],
                file_size,
            ),
        )
        conn.commit()
    finally:
        conn.close()

    return JSONResponse(
        {
            "id": report_id,
            "generatedAt": generated_at.isoformat() + "Z",
            "fileName": file_name,
            "rowCount": result["row_count"],
            "sizeBytes": file_size,
            "downloadUrl": f"/api/reports/{report_id}/download",
        }
    )


@app.get("/api/reports/{report_id}/download")
async def api_download_report(report_id: str):
    conn = get_app_data_db_connection()
    try:
        row = conn.execute(
            "SELECT file_name, file_path FROM report_runs WHERE report_id = ?",
            (report_id,),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return JSONResponse({"message": "Report not found."}, status_code=404)

    file_path = row["file_path"]
    if not os.path.exists(file_path):
        return JSONResponse({"message": "Report file missing on disk."}, status_code=404)

    return FileResponse(
        file_path,
        media_type="application/pdf",
        filename=row["file_name"],
    )


@app.delete("/api/reports/{report_id}")
async def api_delete_report(report_id: str):
    conn = get_app_data_db_connection()
    try:
        row = conn.execute(
            "SELECT file_path FROM report_runs WHERE report_id = ?",
            (report_id,),
        ).fetchone()

        if not row:
            return JSONResponse({"message": "Report not found."}, status_code=404)

        file_path = row["file_path"]

        conn.execute("DELETE FROM report_runs WHERE report_id = ?", (report_id,))
        conn.commit()
    finally:
        conn.close()

    # Delete the PDF file from disk
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except OSError:
            pass  # File already gone or locked – DB row is already removed

    return JSONResponse({"message": "Report deleted successfully.", "id": report_id})

ATTACK_TYPE_COLORS = [
    "#b07cf8",  # purple
    "#00d4ff",  # cyan
    "#f56582",  # pink
    "#f5a623",  # orange
    "#2ee6a8",  # green
    "#a78bfa",  # lavender
    "#00d4ff",  # cyan alt
    "#ef4444",  # red
    "#ec4899",  # pink
    "#06b6d4",  # cyan
]


@app.get("/api/vuln-charts")
async def api_vuln_charts():
    """Return attack-frequency, attack-share, and attack-timeline data from alerts.xlsx."""
    try:
        alerts = alerts_cache.get_alerts()
    except FileNotFoundError:
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    # --- Attack frequency (bar chart) ---
    type_counter = Counter(a.get("type") or "Unknown" for a in alerts)
    attack_frequency = []
    for idx, (attack_type, count) in enumerate(type_counter.most_common(10)):
        color = ATTACK_TYPE_COLORS[idx % len(ATTACK_TYPE_COLORS)]
        attack_frequency.append({"type": attack_type, "count": count, "color": color})

    # --- Attack share (doughnut chart) – same data, all types ---
    attack_share = []
    for idx, (attack_type, count) in enumerate(type_counter.most_common()):
        color = ATTACK_TYPE_COLORS[idx % len(ATTACK_TYPE_COLORS)]
        attack_share.append({"name": attack_type, "value": count, "color": color})

    # --- Attack timeline (line chart – 2-min buckets) ---
    parsed_times = []
    for alert in alerts:
        dt = parse_detected_at(alert.get("detectedAt"))
        if dt:
            parsed_times.append(dt)

    attack_timeline = []
    if parsed_times:
        parsed_times.sort()
        min_time = parsed_times[0].replace(second=0, microsecond=0)
        max_time = parsed_times[-1].replace(second=0, microsecond=0) + timedelta(minutes=2)

        # Align min_time to an even minute
        if min_time.minute % 2 != 0:
            min_time = min_time.replace(minute=min_time.minute - 1)

        bucket_start = min_time
        while bucket_start <= max_time:
            bucket_end = bucket_start + timedelta(minutes=2)
            count = sum(1 for t in parsed_times if bucket_start <= t < bucket_end)
            attack_timeline.append({
                "time": bucket_start.strftime("%H:%M"),
                "count": count,
            })
            bucket_start = bucket_end

    return JSONResponse({
        "attackFrequency": attack_frequency,
        "attackShare": attack_share,
        "attackTimeline": attack_timeline,
        "updatedAt": datetime.utcnow().isoformat() + "Z",
    })


@app.get("/api/target-attack-matrix")
async def api_target_attack_matrix():
    """Pivot alerts by target IP vs attack type."""
    try:
        alerts = alerts_cache.get_alerts()
    except FileNotFoundError:
        return JSONResponse(
            {"message": f"Alerts file not found at {ALERTS_XLSX_PATH}."},
            status_code=404,
        )
    except Exception as exc:
        return JSONResponse({"message": str(exc)}, status_code=500)

    # Collect all unique attack types and build target->type->count map
    all_types: set = set()
    target_map: Dict[str, Counter] = {}

    for alert in alerts:
        target = alert.get("target") or "-"
        attack_type = alert.get("type") or "Unknown"
        all_types.add(attack_type)
        if target not in target_map:
            target_map[target] = Counter()
        target_map[target][attack_type] += 1

    # Sort attack types alphabetically for consistent column order
    sorted_types = sorted(all_types)

    # Build rows: one per target IP, with a count for each attack type
    rows = []
    for target in sorted(target_map.keys()):
        row = {"target": target, "total": sum(target_map[target].values())}
        for attack_type in sorted_types:
            row[attack_type] = target_map[target].get(attack_type, 0)
        rows.append(row)

    # Sort by total attacks descending
    rows.sort(key=lambda r: r["total"], reverse=True)

    return JSONResponse({
        "attackTypes": sorted_types,
        "rows": rows,
        "updatedAt": datetime.utcnow().isoformat() + "Z",
    })


def get_senders_db_connection() -> sqlite3.Connection:
    """Alias for backward compat – senders now live in app_data.db."""
    return get_app_data_db_connection()


@app.get("/api/assets")
async def api_get_assets():
    """Return log-sender data from senders.db as asset cards."""
    if not os.path.exists(APP_DATA_DB_PATH):
        return JSONResponse([])

    conn = get_senders_db_connection()
    try:
        rows = conn.execute(
            "SELECT ip, hostname, total_logs, first_seen, last_seen FROM senders ORDER BY last_seen DESC"
        ).fetchall()
    finally:
        conn.close()

    assets = []
    for idx, row in enumerate(rows, start=1):
        ip = row["ip"]
        total_logs = row["total_logs"]
        first_seen = row["first_seen"]
        last_seen = row["last_seen"]

        # Derive a friendly status from recency
        last_dt = parse_detected_at(last_seen)
        if last_dt:
            age_seconds = (datetime.utcnow() - last_dt).total_seconds()
            if age_seconds < 300:       # active within 5 min
                status = "healthy"
            elif age_seconds < 3600:    # active within 1 hour
                status = "warning"
            elif age_seconds < 86400:   # active within 1 day
                status = "at-risk"
            else:
                status = "offline"
            last_seen_relative = to_relative_time(last_dt)
        else:
            status = "offline"
            last_seen_relative = "Unknown"

        first_dt = parse_detected_at(first_seen)
        first_seen_relative = to_relative_time(first_dt) if first_dt else "Unknown"

        # Decide criticality based on volume
        if total_logs > 1000:
            criticality = "critical"
        elif total_logs > 500:
            criticality = "high"
        elif total_logs > 100:
            criticality = "medium"
        else:
            criticality = "low"

        assets.append({
            "id": f"SND-{idx:03d}",
            "name": f"Log Sender {ip}",
            "type": "server",
            "ip": ip,
            "os": f"{total_logs} logs received",
            "status": status,
            "lastScan": last_seen_relative,
            "vulnerabilities": total_logs,
            "criticality": criticality,
        })

    return JSONResponse(assets)


# ---------------------------------------------------------------------------
# Receive-logs endpoint (formerly on port 6000, now inline)
# ---------------------------------------------------------------------------
@app.post("/receive-logs")
async def receive_logs(request: Request):
    return await receive_logs_handler(request)


# ---------------------------------------------------------------------------
# Reset-All endpoint — clears all operational data (requires password auth)
# ---------------------------------------------------------------------------
@app.post("/api/reset-all")
async def reset_all_data(request: Request):
    """
    Full system reset — wipes ALL operational data:
      • app_data.db tables (alerts, threat_status, report_runs, senders, notifications)
      • All PDF reports from disk
      • input.log — truncated to empty
      • alerts.xlsx, alerts_backup.xlsx, excel2.xlsx — cleared to headers-only
      • Leader blockchain (blockchain.json) — reset to genesis block only
      • Worker blockchain (/admin/reset HTTP call) — reset to genesis block only
    Tables and files are preserved; only their contents are cleared.
    The caller must supply the password of the currently-logged-in user.
    """
    try:
        data = await request.json()
    except Exception:
        return JSONResponse(
            {"success": False, "message": "Invalid request body — JSON expected"},
            status_code=400,
        )

    email = (data.get("email") or "").strip().lower()
    confirmation = (data.get("confirmation") or "").strip()

    if confirmation != "permanently delete all data":
        return JSONResponse(
            {"success": False, "message": "Invalid confirmation phrase. Reset aborted."},
            status_code=400,
        )

    result_details: dict = {}

    # ── 1. Wipe all tables in app_data.db ────────────────────────────────────
    deleted_rows: dict = {}
    conn_data = None
    try:
        conn_data = get_app_data_db_connection()
        tables_to_clear = ["alerts", "threat_status", "report_runs", "senders", "notifications"]
        for table in tables_to_clear:
            try:
                cursor = conn_data.execute(f"DELETE FROM {table}")
                deleted_rows[table] = cursor.rowcount
            except Exception as exc:
                logger.warning("reset-all: could not clear table %s: %s", table, exc)
        conn_data.commit()
        alerts_cache.invalidate()
        logger.info("reset-all: cleared tables %s by user %s", list(deleted_rows.keys()), email)
    except Exception as exc:
        logger.error("reset-all: app_data.db wipe failed: %s", exc)
        return JSONResponse(
            {"success": False, "message": "Failed to clear database tables"},
            status_code=500,
        )
    finally:
        if conn_data:
            conn_data.close()

    result_details["tablesCleared"] = deleted_rows

    # ── 2. Delete all PDF reports from disk ──────────────────────────────────
    deleted_pdfs: list = []
    skipped_pdfs: list = []
    try:
        if os.path.isdir(REPORTS_DIR):
            for fname in os.listdir(REPORTS_DIR):
                if fname.lower().endswith(".pdf"):
                    fpath = os.path.join(REPORTS_DIR, fname)
                    try:
                        os.remove(fpath)
                        deleted_pdfs.append(fname)
                    except Exception as exc:
                        logger.warning("reset-all: could not delete %s: %s", fpath, exc)
                        skipped_pdfs.append(fname)
        logger.info("reset-all: deleted %d PDFs, skipped %d", len(deleted_pdfs), len(skipped_pdfs))
    except Exception as exc:
        logger.error("reset-all: PDF cleanup error: %s", exc)
    result_details["pdfsDeleted"] = len(deleted_pdfs)
    result_details["pdfsSkipped"] = len(skipped_pdfs)

    # ── 3. Truncate input.log ─────────────────────────────────────────────────
    input_log_path = os.path.join(os.path.dirname(__file__), "input.log")
    try:
        with open(input_log_path, "w", encoding="utf-8") as _lf:
            _lf.write("")
        logger.info("reset-all: input.log truncated")
        result_details["inputLogCleared"] = True
    except Exception as exc:
        logger.warning("reset-all: could not truncate input.log: %s", exc)
        result_details["inputLogCleared"] = False

    # ── 4. Clear alerts Excel files (keep headers, remove data rows) ──────────
    import pandas as _pd
    leader_dir = os.path.join(os.path.dirname(__file__), "Blockchain", "leader")
    excel_files_to_clear = {
        "alerts.xlsx": os.path.join(leader_dir, "alerts.xlsx"),
        "alerts_backup.xlsx": os.path.join(leader_dir, "alerts_backup.xlsx"),
        "excel2.xlsx": os.path.join(leader_dir, "excel2.xlsx"),
    }
    excel_results: dict = {}
    for xname, xpath in excel_files_to_clear.items():
        try:
            if os.path.exists(xpath):
                _df_existing = _pd.read_excel(xpath)
                _pd.DataFrame(columns=_df_existing.columns).to_excel(xpath, index=False)
                logger.info("reset-all: cleared %s (had %d rows)", xname, len(_df_existing))
                excel_results[xname] = "cleared"
            else:
                excel_results[xname] = "not found"
        except Exception as exc:
            logger.warning("reset-all: could not clear %s: %s", xname, exc)
            excel_results[xname] = f"error: {exc}"
    result_details["excelFiles"] = excel_results

    # ── 5. Reset leader blockchain to genesis ─────────────────────────────────
    try:
        _leader_module.blockchain.reset_chain()
        logger.info("reset-all: leader blockchain reset to genesis")
        result_details["leaderBlockchainReset"] = True
    except Exception as exc:
        logger.warning("reset-all: could not reset leader blockchain: %s", exc)
        result_details["leaderBlockchainReset"] = False

    # ── 6. Reset worker blockchain(s) via HTTP ────────────────────────────────
    import requests as _req
    worker_reset_results: dict = {}
    try:
        _workers = _leader_module.worker_manager.registry.get("workers", [])
        for _w in _workers:
            _wid = _w.get("id", "?")
            _wurl = f"http://127.0.0.1:{_w.get('port', 5001)}/admin/reset"
            try:
                _wresp = _req.post(_wurl, timeout=4)
                worker_reset_results[f"worker{_wid}"] = _wresp.status_code
                logger.info("reset-all: worker %s reset → %d", _wid, _wresp.status_code)
            except Exception as exc:
                logger.warning("reset-all: worker %s reset error: %s", _wid, exc)
                worker_reset_results[f"worker{_wid}"] = f"unreachable: {exc}"
    except Exception as exc:
        logger.warning("reset-all: worker reset loop failed: %s", exc)
    result_details["workerBlockchainsReset"] = worker_reset_results

    # ── Push SSE so connected UIs refresh immediately ─────────────────────────
    event_bus.emit_threadsafe("alerts.changed", {"trigger": "reset_all"})

    return JSONResponse({
        "success": True,
        "message": "Complete system reset successful. All data has been cleared.",
        "details": result_details,
    })

# ---------------------------------------------------------------------------
# Environment Variables Management
# ---------------------------------------------------------------------------
ENV_FILE_PATH = os.path.join(os.path.dirname(__file__), ".env")


def _read_env_file() -> list:
    """Parse the .env file and return a list of dicts with key, value, and comment info."""
    entries = []
    if not os.path.exists(ENV_FILE_PATH):
        return entries
    with open(ENV_FILE_PATH, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.rstrip("\n").rstrip("\r")
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            if "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            key = key.strip()
            value = value.strip()
            entries.append({"key": key, "value": value})
    return entries


def _write_env_file(entries: list) -> None:
    """Write a list of {key, value} dicts back to the .env file."""
    lines = []
    for entry in entries:
        key = entry.get("key", "").strip()
        value = entry.get("value", "").strip()
        if key:
            lines.append(f"{key}={value}")
    with open(ENV_FILE_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")


@app.get("/api/env")
async def get_env_variables():
    """Return all environment variables from the backend .env file."""
    try:
        entries = _read_env_file()
        return JSONResponse({"success": True, "variables": entries})
    except Exception as exc:
        logger.error("Failed to read .env file: %s", exc)
        return JSONResponse(
            {"success": False, "error": str(exc)}, status_code=500
        )


@app.put("/api/env")
async def update_env_variables(request: Request):
    """Update the backend .env file with the provided variables list."""
    try:
        body = await request.json()
        variables = body.get("variables", [])
        if not isinstance(variables, list):
            return JSONResponse(
                {"success": False, "error": "variables must be a list"},
                status_code=400,
            )
        # Validate each entry has a key
        for entry in variables:
            if not entry.get("key", "").strip():
                return JSONResponse(
                    {"success": False, "error": "Each variable must have a non-empty key"},
                    status_code=400,
                )
        _write_env_file(variables)
        # Re-load the env vars into os.environ so they take effect
        for entry in variables:
            os.environ[entry["key"].strip()] = entry.get("value", "").strip()
        logger.info("Environment variables updated via API (%d entries)", len(variables))
        return JSONResponse({"success": True, "message": f"Updated {len(variables)} variables"})
    except Exception as exc:
        logger.error("Failed to update .env file: %s", exc)
        return JSONResponse(
            {"success": False, "error": str(exc)}, status_code=500
        )


# ---------------------------------------------------------------------------
# SSE event stream for real-time frontend updates
# ---------------------------------------------------------------------------
@app.get("/api/events")
async def sse_event_stream(request: Request):
    """Server-Sent Events endpoint. Frontend connects once and receives
    typed events (alerts.changed, logs.received, blockchain.synced)
    instead of polling."""
    async def _generator():
        async for payload in event_bus.subscribe("*"):
            if await request.is_disconnected():
                break
            yield {
                "event": payload["event"],
                "data": payload.get("data") or "",
            }

    return EventSourceResponse(_generator())


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app:app", host="127.0.0.1", port=8000, reload=True)